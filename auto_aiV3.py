import os
import io
import json
import re
import time
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from google import genai
from google.genai import types

# ✅ 讀取金鑰（本機用 .env，Streamlit Cloud 用 secrets）
from dotenv import load_dotenv
load_dotenv()

API_KEY = os.getenv("GEMINI_API_KEY") or st.secrets.get("GEMINI_API_KEY", "")
if not API_KEY:
    st.error("❌ 找不到 GEMINI_API_KEY，請先設定後重新執行")
    st.stop()

client = genai.Client(api_key=API_KEY.strip())
MODEL = "gemini-3.6-flash"

# 原始位元組小於此值就走 inline，否則走 Files API。
# Gemini inline 請求上限約 20MB，base64 會膨脹約 33%，故取 15MB 保守值。
INLINE_LIMIT = 15 * 1024 * 1024

COLUMN_ORDER = [
    "英文名字＋英文姓氏", "中文姓名", "性別", "國籍", "學歷＋學校＋科系",
    "申請資格條件", "出生日期", "護照號碼", "子領域", "現職公司", "現職職稱",
    "其他工作經歷", "現職是否為主管", "教育背景(學校)", "教育背景(系所)",
    "工作經歷", "產業實績專長", "求學期間", "畢業年份", "總工作年資",
    "審查意見或備注", "勞動部檢核結果", "是否為轉自其他領域", "第1次申請", "年齡", "月薪"
]

OCR_PROMPT = """角色：你是精準的資料分析助理，專責從申請人資料 PDF 中提取關鍵資訊。

任務：
1. 解析 PDF，內容含申請書、護照、學歷、經歷與薪資證明，可能含中文、英文或其他外文。
2. 對每位申請人單獨輸出 JSON（多位則依序生成多筆）。
3. 所有欄位名稱須與指定完全一致，缺漏值填 "N/A"，金額須包含幣別。

工作經歷判斷：
- 若申請為「具數位經濟相關產業、8年以上經驗」：
  1. 判斷公司是否屬軟體/資訊服務業（如軟體開發、系統整合、雲端服務）。
  2. 非軟體業時，須檢視申請人「職務內容」是否符合軟體開發、系統架構或資訊技術核心職能。
  3. 教育/培訓領域不視為數位產業，僅列為輔助經歷。若描述過於簡略，須要求補充專案與貢獻。

輸出格式（僅輸出 JSON 陣列）：
[
  {
    "英文名字＋英文姓氏": "string",
    "中文姓名": "string",
    "性別": "string",
    "國籍": "string",
    "學歷＋學校＋科系": "string，依大學、研究所、博士順序分點",
    "申請資格條件": "string",
    "出生日期": "YYYY-MM-DD",
    "護照號碼": "string",
    "子領域": "string",
    "現職公司": "string",
    "現職職稱": "string",
    "其他工作經歷": "string",
    "現職是否為主管": "string",
    "教育背景(學校)": "string，依大學、研究所、博士順序，格式：大學：校名、研究所：校名",
    "教育背景(系所)": "string，依大學、研究所、博士順序，格式：大學：系所、研究所：系所",
    "工作經歷": "string，列公司、職稱、期間(YYYY-MM)、工作描述與薪資（含幣別）",
    "產業實績專長": "string",
    "求學期間": "string",
    "畢業年份": "數字",
    "總工作年資": "數字，單位年",
    "審查意見或備注": "string，先判定「通過/未通過/待確認」，再詳述評估與疑點（如非軟體業、僅數位工具使用、或描述不足）。",
    "勞動部檢核結果": "string",
    "是否為轉自其他領域": "string",
    "第1次申請": "string",
    "年齡": "string",
    "月薪": "string，含幣別與金額（例：USD 3000，附台幣換算）"
  }
]
"""


def wait_until_active(file_ref, timeout=300, poll=2):
    """Files API 上傳後需等待伺服器處理完成才能用於推論。"""
    deadline = time.time() + timeout
    while file_ref.state.name == "PROCESSING":
        if time.time() > deadline:
            raise TimeoutError(f"檔案處理逾時（{timeout}s）")
        time.sleep(poll)
        file_ref = client.files.get(name=file_ref.name)
    if file_ref.state.name == "FAILED":
        raise RuntimeError("Gemini 檔案處理失敗")
    return file_ref


def extract(file):
    """回傳 (ocr_text, 使用的傳輸方式)。小檔走 inline，大檔走 Files API。"""
    data = file.getvalue()
    mime = file.type or "application/pdf"
    cfg = types.GenerateContentConfig(response_mime_type="application/json")

    if len(data) <= INLINE_LIMIT:
        resp = client.models.generate_content(
            model=MODEL,
            contents=[types.Part.from_bytes(data=data, mime_type=mime), OCR_PROMPT],
            config=cfg,
        )
        return (resp.text or "").strip(), "inline"

    uploaded = client.files.upload(
        file=io.BytesIO(data),
        config=types.UploadFileConfig(mime_type=mime, display_name=file.name),
    )
    try:
        uploaded = wait_until_active(uploaded)
        resp = client.models.generate_content(
            model=MODEL, contents=[uploaded, OCR_PROMPT], config=cfg
        )
        return (resp.text or "").strip(), "files-api"
    finally:
        # 免費配額有總量上限，處理完就清掉
        try:
            client.files.delete(name=uploaded.name)
        except Exception:
            pass


# ✅ Streamlit UI
st.title("📂 金卡檔案 OCR → Excel 轉換工具")

uploaded_files = st.file_uploader(
    "請上傳多個 PDF 或影像檔 (PDF, PNG, JPG)",
    type=["pdf", "png", "jpg", "jpeg"],
    accept_multiple_files=True,
)

if uploaded_files:
    all_records = []
    progress = st.progress(0.0)

    for i, file in enumerate(uploaded_files, start=1):
        size_mb = len(file.getvalue()) / 1024 / 1024
        st.write(f"🔍 處理檔案：**{file.name}**（{size_mb:.1f} MB）...")
        try:
            with st.spinner(f"解析 {file.name} 中..."):
                ocr_text, mode = extract(file)

            if mode == "files-api":
                st.caption("（檔案較大，已改用 Files API 上傳）")

            # response_mime_type 已保證是純 JSON，這兩行僅作保險
            ocr_text = re.sub(r"^```(?:json)?\s*", "", ocr_text)
            ocr_text = re.sub(r"```$", "", ocr_text).replace("﻿", "").strip()

            with st.expander(f"📜 Gemini 回傳內容 ({file.name})"):
                st.code(ocr_text, language="json")

            if not ocr_text:
                st.error(f"❌ {file.name} OCR 失敗：Gemini 沒有回傳內容")
                continue

            try:
                records = json.loads(ocr_text)
            except json.JSONDecodeError:
                st.error(f"❌ {file.name} 回傳不是有效的 JSON，請展開上方內容檢查")
                continue

            if isinstance(records, dict):
                records = [records]
            if isinstance(records, list):
                for r in records:
                    r["來源檔案"] = file.name
                all_records.extend(records)
                st.success(f"✅ {file.name} 解析完成，共 {len(records)} 筆資料")
            else:
                st.warning(f"⚠️ {file.name} 的回傳不是 JSON 陣列格式")

        except Exception as e:
            st.error(f"❌ {file.name} 處理失敗：{e}")
        finally:
            progress.progress(i / len(uploaded_files))

    # 匯總結果轉 Excel
    if all_records:
        df = pd.DataFrame(all_records)
        df = df[[col for col in COLUMN_ORDER if col in df.columns] + ["來源檔案"]]
        st.success(f"🎉 所有檔案完成，共解析 {len(df)} 筆資料")

        buffer = io.BytesIO()
        df.to_excel(buffer, index=False, engine="openpyxl")
        buffer.seek(0)

        # 🔧 openpyxl 美化
        wb = load_workbook(buffer)
        ws = wb.active
        default_font = Font(size=14)

        # 1️⃣ 標題列
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=14)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 2️⃣ 欄寬
        width_map = {
            "英文名字＋英文姓氏": 25, "中文姓名": 12, "性別": 10, "國籍": 12,
            "學歷＋學校＋科系": 30, "申請資格條件": 25, "出生日期": 15, "護照號碼": 18,
            "子領域": 20, "現職公司": 25, "現職職稱": 20, "其他工作經歷": 40,
            "現職是否為主管": 15, "教育背景(學校)": 35, "教育背景(系所)": 35,
            "工作經歷": 80, "產業實績專長": 60, "求學期間": 20, "畢業年份": 12,
            "總工作年資": 15, "審查意見或備注": 80, "勞動部檢核結果": 40,
            "是否為轉自其他領域": 30, "第1次申請": 15, "年齡": 10, "月薪": 25,
            "來源檔案": 30,
        }
        for idx, col_name in enumerate(df.columns, start=1):
            if col_name in width_map:
                col_letter = ws.cell(row=1, column=idx).column_letter
                ws.column_dimensions[col_letter].width = width_map[col_name]

        # 3️⃣ 列高
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 180

        # 4️⃣ 字體、換行、靠上對齊
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.value:
                    cell.font = default_font
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        st.download_button(
            label="⬇️ 下載匯總 Excel",
            data=buffer,
            file_name=f"OCR結果_{now}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.subheader("📊 Excel 預覽")
        st.dataframe(df, use_container_width=True)
